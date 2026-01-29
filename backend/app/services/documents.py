import csv
from pathlib import Path
from typing import Iterable
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import select
from app.core.config import settings
from app.models.models import Document, DocumentChunk
from app.services.embeddings import embed_texts


def _read_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _read_docx(path: Path) -> str:
    doc = DocxDocument(str(path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text)


def _read_csv(path: Path) -> str:
    rows = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            rows.append(" | ".join(cell.strip() for cell in row if cell is not None))
    return "\n".join(rows)


def _read_xlsx(path: Path) -> str:
    workbook = load_workbook(filename=str(path), data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def extract_text(path: Path, file_type: str) -> str:
    file_type = file_type.lower()
    if file_type == "pdf":
        return _read_pdf(path)
    if file_type == "docx":
        return _read_docx(path)
    if file_type == "csv":
        return _read_csv(path)
    if file_type == "xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported document type: {file_type}")


def chunk_text(text: str, max_chars: int = 1200, overlap: int = 200) -> Iterable[str]:
    cleaned = " ".join(text.split())
    if not cleaned:
        return []
    chunks = []
    start = 0
    length = len(cleaned)
    while start < length:
        end = min(start + max_chars, length)
        chunks.append(cleaned[start:end])
        if end == length:
            break
        start = max(0, end - overlap)
    return chunks


def ingest_document(db: Session, document_id: int) -> int:
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise ValueError("Document not found.")
    file_path = Path(settings.document_storage_path) / document.storage_path
    text = extract_text(file_path, document.file_type)
    max_chars = document.chunk_size or 1200
    chunks = list(chunk_text(text, max_chars=max_chars))
    if not chunks:
        return 0
    embeddings = embed_texts(chunks, model=document.embedding_model)
    db.query(DocumentChunk).filter(
        DocumentChunk.document_id == document.id,
        DocumentChunk.company_id == document.company_id,
    ).delete()
    for idx, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
        db.add(DocumentChunk(
            document_id=document.id,
            company_id=document.company_id,
            chunk_index=idx,
            content=chunk,
            embedding=embedding,
        ))
    db.commit()
    return len(chunks)


def search_document_chunks(
    db: Session,
    company_id: int,
    query: str,
    limit: int = 5,
    include_score: bool = False,
) -> list[dict]:
    model_rows = db.query(Document.embedding_model).filter(
        Document.company_id == company_id
    ).distinct().all()
    models = [row[0] for row in model_rows] or [settings.embedding_model]
    all_matches: list[dict] = []
    for model in models:
        query_model = model or settings.embedding_model
        embeddings = embed_texts([query], model=query_model)
        query_vector = embeddings[0]
        distance = DocumentChunk.embedding.cosine_distance(query_vector)
        stmt = (
            select(DocumentChunk, Document, distance)
            .join(Document, Document.id == DocumentChunk.document_id)
            .where(Document.company_id == company_id)
            .where(DocumentChunk.company_id == company_id)
            .where(Document.embedding_model == model)
            .order_by(distance)
            .limit(limit)
        )
        results = db.execute(stmt).all()
        for chunk, document, distance_value in results:
            record = {
                "document_id": document.id,
                "filename": document.filename,
                "file_type": document.file_type,
                "content": chunk.content,
                "chunk_id": chunk.id,
                "embedding_model": query_model,
                "distance": float(distance_value or 0),
            }
            all_matches.append(record)
    all_matches.sort(key=lambda item: item["distance"])
    trimmed = all_matches[:limit]
    if include_score:
        return trimmed
    for item in trimmed:
        item.pop("distance", None)
        item.pop("embedding_model", None)
    return trimmed
